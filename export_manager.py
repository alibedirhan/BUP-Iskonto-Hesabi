# export_manager.py - ÇOKLU PDF DESTEĞİ İLE GÜNCELLENMİŞ
import os
from datetime import datetime
import pandas as pd
from fpdf import FPDF
import tkinter as tk
from tkinter import filedialog, messagebox
import logging
from pathlib import Path

class ExportManager:
    def __init__(self):
        self.setup_directories()
        self.setup_logging()
    
    def setup_logging(self):
        """Loglama ayarları"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
    
    def setup_directories(self):
        """Gerekli dizinleri oluşturur"""
        directories = ['exports', 'exports/excel', 'exports/pdf']
        for directory in directories:
            Path(directory).mkdir(parents=True, exist_ok=True)
            logging.info(f"Dizin hazır: {directory}")
    
    def safe_turkish_text(self, text):
        """Türkçe karakterleri güvenli hale getir - PDF için"""
        if not isinstance(text, str):
            return str(text)
            
        # PDF için ASCII dönüşümü
        replacements = {
            'İ': 'I', 'ı': 'i', 'Ğ': 'G', 'ğ': 'g',
            'Ü': 'U', 'ü': 'u', 'Ş': 'S', 'ş': 's',
            'Ö': 'O', 'ö': 'o', 'Ç': 'C', 'ç': 'c'
        }
        for turkish, ascii_char in replacements.items():
            text = text.replace(turkish, ascii_char)
        return text
    
    def get_clean_filename(self, pdf_name):
        """PDF adından temiz dosya adı oluştur"""
        # Uzantıyı kaldır
        base_name = pdf_name.replace('.pdf', '').replace('.PDF', '')
        
        # Özel karakterleri temizle
        clean_name = "".join(c for c in base_name if c.isalnum() or c in (' ', '-', '_'))
        
        # Fazla boşlukları temizle
        clean_name = ' '.join(clean_name.split())
        
        return clean_name
    
    def export_to_excel_multi(self, all_pdf_data, discount_vars):
        """Çoklu PDF'i Excel'e aktar - Her PDF için ayrı sheet"""
        try:
            if not all_pdf_data:
                messagebox.showwarning("Uyarı", "Dışa aktarılacak veri bulunamadı!")
                return
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            initial_filename = f"Bupilic_Iskontolu_Fiyat_Listeleri_{timestamp}.xlsx"
            
            file_path = filedialog.asksaveasfilename(
                title="Excel olarak kaydet",
                defaultextension=".xlsx",
                initialfile=initial_filename,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Önce özet sayfası
                summary_data = self._create_multi_summary(all_pdf_data, discount_vars)
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='ÖZET', index=False)
                
                # Her PDF için ayrı sheet
                for sheet_idx, (pdf_name, pdf_data) in enumerate(all_pdf_data.items(), 1):
                    # Sheet adını oluştur (Excel 31 karakter sınırı)
                    sheet_name = f"{sheet_idx}_{self.get_clean_filename(pdf_name)}"[:31]
                    
                    # PDF verilerini organize et
                    sheet_data = []
                    
                    # Başlık satırı
                    sheet_data.append({
                        'Kategori': f"PDF: {pdf_name}",
                        'Sıra': '',
                        'Ürün Adı': '',
                        'Orj. KDV Hariç (TL)': '',
                        'Orj. KDV Dahil (TL)': '',
                        'İsk. KDV Hariç (TL)': '',
                        'İsk. KDV Dahil (TL)': '',
                        'İskonto Oranı (%)': '',
                        'İskonto Tutarı (TL)': ''
                    })
                    
                    # Boş satır
                    sheet_data.append({col: '' for col in sheet_data[0].keys()})
                    
                    # Her kategori için veriler
                    for category, products in pdf_data['data'].items():
                        if not products:
                            continue
                        
                        discount_rate = discount_vars[category].get()
                        
                        # Kategori başlığı
                        sheet_data.append({
                            'Kategori': f"{category} - %{discount_rate:.1f} İSKONTO",
                            'Sıra': '',
                            'Ürün Adı': '',
                            'Orj. KDV Hariç (TL)': '',
                            'Orj. KDV Dahil (TL)': '',
                            'İsk. KDV Hariç (TL)': '',
                            'İsk. KDV Dahil (TL)': '',
                            'İskonto Oranı (%)': '',
                            'İskonto Tutarı (TL)': ''
                        })
                        
                        # Sütun başlıkları
                        sheet_data.append({
                            'Kategori': 'Kategori',
                            'Sıra': 'Sıra',
                            'Ürün Adı': 'Ürün Adı',
                            'Orj. KDV Hariç (TL)': 'Orj. KDV Hariç (TL)',
                            'Orj. KDV Dahil (TL)': 'Orj. KDV Dahil (TL)',
                            'İsk. KDV Hariç (TL)': 'İsk. KDV Hariç (TL)',
                            'İsk. KDV Dahil (TL)': 'İsk. KDV Dahil (TL)',
                            'İskonto Oranı (%)': 'İskonto Oranı (%)',
                            'İskonto Tutarı (TL)': 'İskonto Tutarı (TL)'
                        })
                        
                        # Ürünler
                        for idx, product in enumerate(products, 1):
                            iskonto_tutari = product.get('original_price_with_vat', 0) - product['price_with_vat']
                            
                            sheet_data.append({
                                'Kategori': category,
                                'Sıra': idx,
                                'Ürün Adı': product['name'],
                                'Orj. KDV Hariç (TL)': product.get('original_price_without_vat', 0),
                                'Orj. KDV Dahil (TL)': product.get('original_price_with_vat', 0),
                                'İsk. KDV Hariç (TL)': product['price_without_vat'],
                                'İsk. KDV Dahil (TL)': product['price_with_vat'],
                                'İskonto Oranı (%)': discount_rate,
                                'İskonto Tutarı (TL)': round(iskonto_tutari, 2)
                            })
                        
                        # Kategoriler arası boş satır
                        sheet_data.append({col: '' for col in sheet_data[0].keys()})
                        sheet_data.append({col: '' for col in sheet_data[0].keys()})
                    
                    # DataFrame oluştur ve Excel'e yaz
                    df = pd.DataFrame(sheet_data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Formatla
                    self._format_excel_sheet(writer, sheet_name, df)
            
            logging.info(f"Çoklu Excel dosyası kaydedildi: {file_path}")
            messagebox.showinfo("Başarılı", 
                              f"Excel dosyası başarıyla kaydedildi:\n{file_path}\n\n"
                              f"{len(all_pdf_data)} PDF için ayrı sheet oluşturuldu.")
            
        except PermissionError:
            messagebox.showerror("Hata", "Dosya başka bir program tarafından kullanılıyor.")
        except Exception as e:
            logging.error(f"Excel kaydetme hatası: {e}")
            messagebox.showerror("Hata", f"Excel kaydetme hatası: {str(e)}")
    
    def export_to_pdf_multi(self, all_pdf_data, discount_vars):
        """Her PDF için ayrı iskontolu PDF dosyası oluştur"""
        try:
            if not all_pdf_data:
                messagebox.showwarning("Uyarı", "Dışa aktarılacak veri bulunamadı!")
                return
            
            # Kayıt klasörünü seç
            save_dir = filedialog.askdirectory(title="PDF'leri kaydetmek için klasör seçin")
            
            if not save_dir:
                return
            
            saved_files = []
            current_date = datetime.now().strftime("%d.%m.%Y")
            
            for pdf_name, pdf_data in all_pdf_data.items():
                # Dosya adını oluştur
                clean_name = self.get_clean_filename(pdf_name)
                output_filename = f"{clean_name}_İskontolu_{current_date}.pdf"
                file_path = os.path.join(save_dir, output_filename)
                
                # PDF oluştur
                pdf = SafePDF()
                pdf.set_auto_page_break(auto=True, margin=15)
                
                # Kapak sayfası
                pdf.add_page()
                self._add_pdf_cover_page(pdf, pdf_name, pdf_data['data'], discount_vars)
                
                # Özet sayfası
                pdf.add_page()
                self._add_pdf_summary_page(pdf, pdf_data['data'], discount_vars)
                
                # Kategori sayfaları
                for category, products in pdf_data['data'].items():
                    if not products:
                        continue
                    
                    discount_rate = discount_vars[category].get()
                    
                    pdf.add_page()
                    safe_category = self.safe_turkish_text(category)
                    self._add_pdf_header(pdf, f"{safe_category} - %{discount_rate:.1f} Iskonto")
                    
                    # Tablo başlıkları
                    pdf.set_font("Arial", 'B', 9)
                    pdf.cell(10, 8, "No", 1, 0, 'C')
                    pdf.cell(80, 8, "Urun Adi", 1, 0, 'C')
                    pdf.cell(25, 8, "Orj.KDV Haric", 1, 0, 'C')
                    pdf.cell(25, 8, "Orj.KDV Dahil", 1, 0, 'C')
                    pdf.cell(25, 8, "Isk.KDV Haric", 1, 0, 'C')
                    pdf.cell(25, 8, "Isk.KDV Dahil", 1, 1, 'C')
                    
                    pdf.set_font("Arial", '', 8)
                    
                    for idx, product in enumerate(products, 1):
                        # Yeni sayfa kontrolü
                        if pdf.get_y() > 260:
                            pdf.add_page()
                            self._add_pdf_header(pdf, f"{safe_category} - %{discount_rate:.1f} Iskonto (devam)")
                            
                            # Başlıkları tekrarla
                            pdf.set_font("Arial", 'B', 9)
                            pdf.cell(10, 8, "No", 1, 0, 'C')
                            pdf.cell(80, 8, "Urun Adi", 1, 0, 'C')
                            pdf.cell(25, 8, "Orj.KDV Haric", 1, 0, 'C')
                            pdf.cell(25, 8, "Orj.KDV Dahil", 1, 0, 'C')
                            pdf.cell(25, 8, "Isk.KDV Haric", 1, 0, 'C')
                            pdf.cell(25, 8, "Isk.KDV Dahil", 1, 1, 'C')
                            pdf.set_font("Arial", '', 8)
                        
                        # Ürün adını uygun uzunluğa getir
                        name = self.safe_turkish_text(product['name'])
                        if len(name) > 45:
                            name = name[:42] + "..."
                        
                        pdf.cell(10, 7, str(idx), 1, 0, 'C')
                        pdf.cell(80, 7, name, 1, 0, 'L')
                        pdf.cell(25, 7, f"{product.get('original_price_without_vat', 0):.2f}", 1, 0, 'R')
                        pdf.cell(25, 7, f"{product.get('original_price_with_vat', 0):.2f}", 1, 0, 'R')
                        pdf.cell(25, 7, f"{product['price_without_vat']:.2f}", 1, 0, 'R')
                        pdf.cell(25, 7, f"{product['price_with_vat']:.2f}", 1, 1, 'R')
                
                # PDF'i kaydet
                pdf.output(file_path)
                saved_files.append(output_filename)
                logging.info(f"PDF kaydedildi: {file_path}")
            
            # Başarı mesajı
            messagebox.showinfo("Başarılı", 
                              f"{len(saved_files)} PDF dosyası başarıyla kaydedildi:\n\n" + 
                              "\n".join(saved_files) +
                              f"\n\nKayıt yeri: {save_dir}")
            
        except Exception as e:
            logging.error(f"PDF kaydetme hatası: {e}")
            messagebox.showerror("Hata", f"PDF kaydetme hatası: {str(e)}")
    
    def _create_multi_summary(self, all_pdf_data, discount_vars):
        """Çoklu PDF için özet oluştur"""
        summary = []
        
        for pdf_idx, (pdf_name, pdf_data) in enumerate(all_pdf_data.items(), 1):
            pdf_summary = {
                'PDF No': pdf_idx,
                'PDF Adı': pdf_name,
                'Toplam Ürün': 0,
                'Toplam Orijinal (TL)': 0,
                'Toplam İskontolu (TL)': 0,
                'Toplam İskonto (TL)': 0
            }
            
            for category, products in pdf_data['data'].items():
                if products:
                    pdf_summary['Toplam Ürün'] += len(products)
                    
                    for product in products:
                        orig = product.get('original_price_with_vat', 0)
                        disc = product['price_with_vat']
                        pdf_summary['Toplam Orijinal (TL)'] += orig
                        pdf_summary['Toplam İskontolu (TL)'] += disc
                        pdf_summary['Toplam İskonto (TL)'] += (orig - disc)
            
            pdf_summary['Toplam Orijinal (TL)'] = round(pdf_summary['Toplam Orijinal (TL)'], 2)
            pdf_summary['Toplam İskontolu (TL)'] = round(pdf_summary['Toplam İskontolu (TL)'], 2)
            pdf_summary['Toplam İskonto (TL)'] = round(pdf_summary['Toplam İskonto (TL)'], 2)
            
            summary.append(pdf_summary)
        
        return summary
    
    def _format_excel_sheet(self, writer, sheet_name, df):
        """Excel sheet'ini formatla"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        worksheet = writer.sheets[sheet_name]
        
        # Sütun genişlikleri
        column_widths = {
            'A': 25,  # Kategori
            'B': 8,   # Sıra
            'C': 50,  # Ürün Adı
            'D': 18,  # Orj. KDV Hariç
            'E': 18,  # Orj. KDV Dahil
            'F': 18,  # İsk. KDV Hariç
            'G': 18,  # İsk. KDV Dahil
            'H': 15,  # İskonto Oranı
            'I': 18   # İskonto Tutarı
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
    
    def _add_pdf_cover_page(self, pdf, pdf_name, data, discount_vars):
        """PDF kapak sayfası"""
        pdf.set_font("Arial", 'B', 24)
        pdf.ln(50)
        pdf.cell(0, 15, "BUPILIC", 0, 1, 'C')
        pdf.set_font("Arial", '', 18)
        pdf.cell(0, 10, "ISKONTOLU FIYAT LISTESI", 0, 1, 'C')
        
        pdf.ln(20)
        pdf.set_font("Arial", '', 14)
        clean_name = self.safe_turkish_text(self.get_clean_filename(pdf_name))
        pdf.cell(0, 10, clean_name, 0, 1, 'C')
        
        pdf.ln(10)
        pdf.set_font("Arial", '', 12)
        pdf.cell(0, 10, f"Olusturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", 0, 1, 'C')
        
        # Özet bilgiler
        pdf.ln(30)
        total_products = sum(len(products) for products in data.values() if products)
        pdf.set_font("Arial", '', 11)
        pdf.cell(0, 8, f"Toplam Urun Sayisi: {total_products}", 0, 1, 'C')
        
        categories_with_products = [cat for cat, products in data.items() if products]
        pdf.cell(0, 8, f"Kategori Sayisi: {len(categories_with_products)}", 0, 1, 'C')
    
    def _add_pdf_summary_page(self, pdf, data, discount_vars):
        """PDF özet sayfası"""
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(0, 10, "OZET BILGILER", 0, 1, 'L')
        pdf.ln(5)
        
        # Özet tablo
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(60, 8, "Kategori", 1, 0, 'C')
        pdf.cell(25, 8, "Urun Sayisi", 1, 0, 'C')
        pdf.cell(30, 8, "Iskonto (%)", 1, 0, 'C')
        pdf.cell(35, 8, "Toplam Iskonto", 1, 1, 'C')
        
        pdf.set_font("Arial", '', 9)
        total_products = 0
        total_discount = 0
        
        for category, products in data.items():
            if products:
                count = len(products)
                rate = discount_vars[category].get()
                discount = sum(p.get('original_price_with_vat', 0) - p['price_with_vat'] for p in products)
                
                safe_category = self.safe_turkish_text(category)
                
                pdf.cell(60, 7, safe_category[:30], 1, 0, 'L')
                pdf.cell(25, 7, str(count), 1, 0, 'C')
                pdf.cell(30, 7, f"{rate:.1f}", 1, 0, 'C')
                pdf.cell(35, 7, f"{discount:.2f} TL", 1, 1, 'R')
                
                total_products += count
                total_discount += discount
        
        # Toplam satırı
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(60, 8, "TOPLAM", 1, 0, 'L')
        pdf.cell(25, 8, str(total_products), 1, 0, 'C')
        pdf.cell(30, 8, "-", 1, 0, 'C')
        pdf.cell(35, 8, f"{total_discount:.2f} TL", 1, 1, 'R')
    
    def _add_pdf_header(self, pdf, title):
        """PDF başlık ekler"""
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "BUPILIC ISKONTOLU FIYAT LISTESI", 0, 1, 'C')
        
        pdf.set_font("Arial", '', 12)
        pdf.cell(0, 8, title, 0, 1, 'C')
        
        pdf.set_font("Arial", '', 10)
        pdf.cell(0, 8, f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}", 0, 1, 'C')
        pdf.ln(5)
    
    # Eski single PDF metodları (geriye uyumluluk için)
    def export_to_excel(self, data, discount_vars):
        """Tek PDF için Excel export - geriye uyumluluk"""
        all_data = {"Fiyat Listesi": {"data": data, "type": "normal", "path": ""}}
        self.export_to_excel_multi(all_data, discount_vars)
    
    def export_to_pdf(self, data, discount_vars):
        """Tek PDF için PDF export - geriye uyumluluk"""
        all_data = {"Fiyat_Listesi": {"data": data, "type": "normal", "path": ""}}
        self.export_to_pdf_multi(all_data, discount_vars)


class SafePDF(FPDF):
    """Güvenli UTF-8 destekli PDF sınıfı"""
    def __init__(self):
        super().__init__()
        self.font_loaded = False
        self._load_fonts()
        
    def _load_fonts(self):
        """Fontları güvenli şekilde yükle"""
        try:
            # Windows font yolları
            windows_paths = [
                'C:/Windows/Fonts/arial.ttf',
                'C:/Windows/Fonts/arialbd.ttf'
            ]
            
            # Linux font yolları
            linux_paths = [
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'
            ]
            
            # Font yüklemeyi dene
            for normal_path, bold_path in [(windows_paths[0], windows_paths[1])] + \
                                         list(zip(linux_paths[::2], linux_paths[1::2])):
                try:
                    if Path(normal_path).exists() and Path(bold_path).exists():
                        self.add_font('Arial', '', normal_path, uni=True)
                        self.add_font('Arial', 'B', bold_path, uni=True)
                        self.font_loaded = True
                        logging.info(f"Font yüklendi: {normal_path}")
                        break
                except Exception:
                    continue
                    
        except Exception as e:
            logging.warning(f"Font yükleme başarısız: {e}")
    
    def set_font(self, family, style='', size=0):
        """Güvenli font ayarlama"""
        try:
            if self.font_loaded and family == 'Arial':
                super().set_font('Arial', style, size)
            else:
                # Fallback font
                super().set_font('Helvetica', style, size)
        except Exception:
            # Son çare
            super().set_font('Helvetica', '', 12)
    
    def header(self):
        """Sayfa başlığı"""
        pass  # Manuel olarak eklenecek
    
    def footer(self):
        """Sayfa alt bilgisi"""
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Sayfa {self.page_no()}', 0, 0, 'C')